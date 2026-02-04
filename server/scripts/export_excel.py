#!/usr/bin/env python3
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def main():
    if len(sys.argv) != 3:
        print("Usage: export_excel.py <input_json> <output_xlsx>", file=sys.stderr)
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    # Leggi dati JSON
    with open(input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingredienti"
    
    # Header
    headers = ["ID", "Nome", "Categoria", "Fornitore", "Unità", "Qtà Confezione", 
               "Prezzo Confezione (€)", "Prezzo/kg o unità (€)", "Marca", "Note", "Food", "Allergeni"]
    ws.append(headers)
    
    # Stile header
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Dati
    for ing in data:
        allergens_str = ", ".join(ing.get("allergens", [])) if ing.get("allergens") else ""
        ws.append([
            ing.get("id", ""),
            ing.get("name", ""),
            ing.get("category", ""),
            ing.get("supplierName", ""),
            "kg" if ing.get("unitType") == "k" else "pz",
            float(ing.get("packageQuantity", 0)),
            float(ing.get("packagePrice", 0)),
            float(ing.get("pricePerKgOrUnit", 0)),
            ing.get("brand", ""),
            ing.get("notes", ""),
            "Sì" if ing.get("isFood") else "No",
            allergens_str
        ])
    
    # Auto-size colonne
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva
    wb.save(output_file)
    print("OK")

if __name__ == "__main__":
    main()
