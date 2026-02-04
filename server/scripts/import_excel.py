#!/usr/bin/env python3
import json
import sys
from openpyxl import load_workbook

def main():
    if len(sys.argv) != 3:
        print("Usage: import_excel.py <input_xlsx> <output_json>", file=sys.stderr)
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    # Carica workbook
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Leggi header (prima riga)
    headers = [cell.value for cell in ws[1]]
    
    # Mappa header a indici
    header_map = {h: i for i, h in enumerate(headers)}
    
    # Leggi dati
    ingredients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[header_map.get("Nome", 1)]:  # Skip righe vuote
            continue
            
        ing = {
            "name": str(row[header_map.get("Nome", 1)] or ""),
            "category": str(row[header_map.get("Categoria", 2)] or "Altro"),
            "supplier": str(row[header_map.get("Fornitore", 3)] or "Non specificato"),
            "unit": str(row[header_map.get("Unità", 4)] or "kg"),
            "packageQuantity": float(row[header_map.get("Qtà Confezione", 5)] or 0),
            "packagePrice": float(row[header_map.get("Prezzo Confezione (€)", 6)] or 0),
            "brand": str(row[header_map.get("Marca", 8)] or ""),
            "notes": str(row[header_map.get("Note", 9)] or ""),
            "isFood": str(row[header_map.get("Food", 10)] or "Sì").lower() in ["sì", "si", "yes", "true", "1"],
            "allergens": str(row[header_map.get("Allergeni", 11)] or "").split(", ") if row[header_map.get("Allergeni", 11)] else []
        }
        
        # Calcola pricePerKgOrUnit
        if ing["packageQuantity"] > 0:
            ing["pricePerKgOrUnit"] = ing["packagePrice"] / ing["packageQuantity"]
        else:
            ing["pricePerKgOrUnit"] = 0
            
        ingredients.append(ing)
    
    # Salva JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(ingredients, f, ensure_ascii=False, indent=2)
    
    print(f"OK: {len(ingredients)} ingredienti")

if __name__ == "__main__":
    main()
