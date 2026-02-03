import openpyxl
import os
from urllib.parse import urlparse
import mysql.connector

# Carica il file Excel
wb = openpyxl.load_workbook('/home/ubuntu/upload/MENUUNION2026.xlsx', data_only=True)
ws = wb['PREZZO_FINALE']

# Connessione database
db_url = os.environ.get('DATABASE_URL', '')
parsed = urlparse(db_url)

conn = mysql.connector.connect(
    host=parsed.hostname,
    port=parsed.port or 3306,
    user=parsed.username,
    password=parsed.password,
    database=parsed.path.lstrip('/'),
    ssl_ca='',
    ssl_verify_cert=False
)
cursor = conn.cursor(dictionary=True)

# Leggi peso unitario dal foglio PREZZO_FINALE
headers = [cell.value for cell in ws[1]]
codice_idx = headers.index('CODICE_RICETTA')
um_idx = headers.index('UM')
peso_unitario_idx = headers.index('Peso unitario')

unit_weights = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[codice_idx]:
        continue
    
    codice = str(row[codice_idx]).strip()
    um = row[um_idx]
    peso_unitario = row[peso_unitario_idx]
    
    # Se ha peso unitario, è venduto a unità (anche se UM=K)
    unit_type = 'u' if peso_unitario else 'k'
    
    # Peso unitario in grammi (se presente)
    unit_weight = None
    if peso_unitario:
        unit_weight = float(peso_unitario)
    
    unit_weights[codice] = {
        'unitType': unit_type,
        'unitWeight': unit_weight
    }

print(f"=== UNIT WEIGHTS TROVATI ===")
print(f"Totale ricette: {len(unit_weights)}")
for code, data in list(unit_weights.items())[:5]:
    print(f"{code}: type={data['unitType']}, weight={data['unitWeight']}")

# Aggiorna il database
print(f"\n=== AGGIORNAMENTO DATABASE ===")
updated_count = 0

for codice, data in unit_weights.items():
    cursor.execute(
        "UPDATE final_recipes SET unitType = %s, unitWeight = %s WHERE code = %s",
        (data['unitType'], data['unitWeight'], codice)
    )
    if cursor.rowcount > 0:
        updated_count += 1
        print(f"✓ {codice}: unitType={data['unitType']}, unitWeight={data['unitWeight']}")

conn.commit()
cursor.close()
conn.close()

print(f"\n=== COMPLETATO ===")
print(f"Ricette aggiornate: {updated_count}/{len(unit_weights)}")

wb.close()
