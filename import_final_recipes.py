#!/usr/bin/env python3
"""
Script per importare ricette finali dal foglio REC_FINALE di MENUUNION2026.xlsx
"""
import openpyxl
import mysql.connector
import os
import json
from urllib.parse import urlparse

# Parse DATABASE_URL
db_url = os.environ.get('DATABASE_URL')
if not db_url:
    print("ERROR: DATABASE_URL not set")
    exit(1)

parsed = urlparse(db_url)
username = parsed.username
password = parsed.password
hostname = parsed.hostname
port = parsed.port or 3306
database = parsed.path.lstrip('/')

# Connessione al database
conn = mysql.connector.connect(
    host=hostname,
    port=port,
    user=username,
    password=password,
    database=database
)
cursor = conn.cursor(dictionary=True)

print("=== IMPORTAZIONE RICETTE FINALI ===\n")

# Carica il file Excel
wb = openpyxl.load_workbook('/home/ubuntu/upload/MENUUNION2026.xlsx')
ws = wb['REC_FINALE']

# Raggruppa componenti per RC_ID
recipes_dict = {}

for row in range(2, ws.max_row + 1):
    rc_id = ws.cell(row, 1).value  # RC_ID
    component_name = ws.cell(row, 2).value  # COMPONENTE
    cod = ws.cell(row, 3).value  # COD (prodotto/ops/semilavorato)
    um = ws.cell(row, 4).value  # UM
    qty = ws.cell(row, 6).value  # Qty
    
    if not rc_id or not component_name:
        continue
    
    # Inizializza ricetta se non esiste
    if rc_id not in recipes_dict:
        recipes_dict[rc_id] = {
            'code': rc_id,
            'name': rc_id.replace('_', ' ').title(),
            'components': []
        }
    
    # Trova l'ingrediente nel database
    cursor.execute("""
        SELECT id, name, pricePerKgOrUnit, unitType 
        FROM ingredients 
        WHERE name = %s
        LIMIT 1
    """, (component_name,))
    
    ingredient = cursor.fetchone()
    
    if ingredient:
        # Gestisci formule Excel
        qty_value = 0
        if qty:
            if isinstance(qty, (int, float)):
                qty_value = float(qty)
            elif isinstance(qty, str) and qty.startswith('='):
                # Salta formule Excel, usa 0 come default
                qty_value = 0
            else:
                try:
                    qty_value = float(qty)
                except:
                    qty_value = 0
        
        recipes_dict[rc_id]['components'].append({
            'ingredientId': ingredient['id'],
            'ingredientName': ingredient['name'],
            'quantity': qty_value,
            'unitType': um or 'g',
            'pricePerUnit': float(ingredient['pricePerKgOrUnit']) if ingredient['pricePerKgOrUnit'] else 0
        })

print(f"Trovate {len(recipes_dict)} ricette finali uniche\n")

# Inserisci ricette nel database
imported_count = 0
skipped_count = 0

for rc_id, recipe_data in recipes_dict.items():
    # Verifica se la ricetta esiste già
    cursor.execute("""
        SELECT id FROM final_recipes WHERE code = %s
    """, (recipe_data['code'],))
    
    existing = cursor.fetchone()
    
    if existing:
        print(f"⚠️  Ricetta {recipe_data['code']} già esistente, aggiorno componenti...")
        # Aggiorna i componenti
        cursor.execute("""
            UPDATE final_recipes 
            SET components = %s
            WHERE code = %s
        """, (json.dumps(recipe_data['components']), recipe_data['code']))
        skipped_count += 1
    else:
        # Inserisci nuova ricetta
        import nanoid
        recipe_id = nanoid.generate(size=21)
        
        # Calcola il costo totale
        total_cost = sum(comp['quantity'] * comp['pricePerUnit'] / 1000 if comp['unitType'] == 'g' else comp['quantity'] * comp['pricePerUnit'] for comp in recipe_data['components'])
        
        cursor.execute("""
            INSERT INTO final_recipes (
                id, code, name, category, yieldPercentage, 
                conservationMethod, maxConservationTime, serviceWastePercentage, 
                totalCost, components, createdAt, updatedAt
            ) VALUES (
                %s, %s, %s, %s, %s, 
                %s, %s, %s, 
                %s, %s, NOW(), NOW()
            )
        """, (
            recipe_id,
            recipe_data['code'],
            recipe_data['name'],
            'Pane',  # Categoria di default
            1.000,  # Resa 100% (formato decimal 5,3)
            'Refrigerato',
            '3 giorni',  # Tempo conservazione
            0.005,  # Scarto servizio 0.5% (formato decimal 5,3)
            round(total_cost, 2),
            json.dumps(recipe_data['components'])
        ))
        
        imported_count += 1
        print(f"✅ Importata ricetta: {recipe_data['name']} ({len(recipe_data['components'])} componenti)")

conn.commit()

print(f"\n=== RIEPILOGO ===")
print(f"Ricette importate: {imported_count}")
print(f"Ricette aggiornate: {skipped_count}")
print(f"Totale ricette: {len(recipes_dict)}")

cursor.close()
conn.close()
