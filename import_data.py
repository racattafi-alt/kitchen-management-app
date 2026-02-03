#!/usr/bin/env python3
"""
Script per importare dati da MENUUNION2026.xlsx nel database Kitchen Management
"""
import openpyxl
import mysql.connector
import os
from nanoid import generate
from decimal import Decimal

# Connessione al database
db_url = os.environ.get('DATABASE_URL', '')
# Parse DATABASE_URL: mysql://user:password@host:port/database?params
if db_url.startswith('mysql://'):
    db_url = db_url.replace('mysql://', '')
    if '@' in db_url:
        auth, host_db = db_url.split('@', 1)
        user, password = auth.split(':', 1)
        # Rimuovi parametri query
        if '?' in host_db:
            host_db = host_db.split('?')[0]
        if '/' in host_db:
            host_port, database = host_db.split('/', 1)
            if ':' in host_port:
                host, port = host_port.split(':', 1)
            else:
                host, port = host_port, '3306'
        else:
            host, port, database = host_db, '3306', 'kitchen'
    else:
        raise ValueError("Invalid DATABASE_URL format")
else:
    raise ValueError("DATABASE_URL must start with mysql://")

conn = mysql.connector.connect(
    host=host,
    port=int(port),
    user=user,
    password=password,
    database=database
)
cursor = conn.cursor()

print("Connesso al database")

# Carica il file Excel
wb = openpyxl.load_workbook('/home/ubuntu/upload/MENUUNION2026.xlsx', data_only=True)

# Mappa categorie Excel -> categorie DB
category_map = {
    'Additivi': 'Additivi',
    'Carni': 'Carni',
    'Latticini e uovo': 'Latticini',
    'General Supplies': 'Altro',
    'Verdura': 'Verdura',
    'Spezie': 'Spezie',
    'Farine': 'Farine',
}

# ============ IMPORTA INGREDIENTI ============
print("\n=== Importazione Ingredienti ===")
ws_prodotti = wb['DB_costo prodotti']
ingredienti_importati = 0

for row in ws_prodotti.iter_rows(min_row=2, values_only=True):
    if not row[0]:  # Skip righe vuote
        continue
    
    nome = row[0]  # Articolo
    unit_type = 'k' if row[1] == 'k' else 'u'
    qty = float(row[2]) if row[2] else 1.0
    fornitore = row[4] if row[4] else 'Non specificato'
    categoria_excel = row[5] if row[5] else 'Altro'
    categoria = category_map.get(categoria_excel, 'Altro')
    
    # Calcola prezzo
    try:
        prezzo_conf = float(row[6]) if isinstance(row[6], (int, float)) else 0.0
        prezzo_per_kg = float(row[7]) if isinstance(row[7], (int, float)) else 0.0
        
        if prezzo_per_kg == 0 and prezzo_conf > 0:
            if unit_type == 'k':
                prezzo_per_kg = (prezzo_conf * 1000) / qty
            else:
                prezzo_per_kg = prezzo_conf / qty
    except:
        prezzo_per_kg = 0.0
    
    # Inserisci nel database
    try:
        cursor.execute("""
            INSERT INTO ingredients (id, name, supplier, category, unitType, packageQuantity, 
                                   packagePrice, pricePerKgOrUnit, brand, notes, isActive)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            generate(size=21),
            nome,
            fornitore,
            categoria,
            unit_type,
            str(qty),
            str(prezzo_conf),
            str(prezzo_per_kg),
            None,
            None,
            True
        ))
        ingredienti_importati += 1
    except mysql.connector.IntegrityError:
        # Ingrediente già esistente, skip
        pass
    except Exception as e:
        print(f"Errore importando {nome}: {e}")

conn.commit()
print(f"Importati {ingredienti_importati} ingredienti")

# ============ IMPORTA SEMILAVORATI ============
print("\n=== Importazione Semilavorati ===")
ws_semi = wb['DB_semilavorati']
ws_rec_semi = wb['REC_semilavorati']
semilavorati_importati = 0

# Crea un dizionario dei semilavorati
semilavorati_dict = {}
for row in ws_semi.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    codice = row[0]
    nome = row[1]
    categoria_excel = row[2]
    
    # Mappa categoria semilavorati
    cat_map_semi = {
        'SPEZIE': 'SPEZIE',
        'SALSE': 'SALSE',
        'VERDURA': 'VERDURA',
        'CARNE': 'CARNE',
    }
    categoria = cat_map_semi.get(categoria_excel, 'ALTRO')
    
    semilavorati_dict[codice] = {
        'nome': nome,
        'categoria': categoria,
        'componenti': []
    }

# Aggiungi componenti ai semilavorati
for row in ws_rec_semi.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    sl_id = row[0]
    ingrediente = row[1]
    qty = float(row[2]) if row[2] else 0.0
    um = float(row[3]) if row[3] else 1000.0
    
    if sl_id in semilavorati_dict:
        semilavorati_dict[sl_id]['componenti'].append({
            'ingredientId': ingrediente,
            'quantity': qty,
            'unitType': 'g'
        })

# Inserisci semilavorati nel database
for codice, data in semilavorati_dict.items():
    try:
        import json
        cursor.execute("""
            INSERT INTO semi_finished_recipes (id, code, name, category, yieldPercentage, 
                                            shelfLifeDays, storageMethod, components, 
                                            productionSteps, finalPricePerKg, totalQuantityProduced)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            generate(size=21),
            codice,
            data['nome'],
            data['categoria'],
            '95.0',  # Default yield
            7,  # Default shelf life
            'Refrigerato',
            json.dumps(data['componenti']),
            json.dumps([]),
            '0.0',  # Sarà calcolato
            None
        ))
        semilavorati_importati += 1
    except mysql.connector.IntegrityError:
        pass
    except Exception as e:
        print(f"Errore importando semilavorato {codice}: {e}")

conn.commit()
print(f"Importati {semilavorati_importati} semilavorati")

# ============ IMPORTA RICETTE FINALI ============
print("\n=== Importazione Ricette Finali ===")
# Usa il foglio DB_semilavorati per le ricette finali
ws_finali = wb['DB_semilavorati']
ws_rec_finali = wb['REC_FINALE']
ricette_importate = 0

# Crea dizionario ricette finali
ricette_dict = {}
for row in ws_finali.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    codice = row[0]
    
    ricette_dict[codice] = {
        'componenti': []
    }

# Aggiungi componenti alle ricette
for row in ws_rec_finali.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rc_id = row[0]
    componente = row[1]
    cod_tipo = row[2]  # prodotto, ops, semilavorato
    um = row[3]
    qty = float(row[5]) if row[5] else 0.0
    
    if rc_id in ricette_dict:
        ricette_dict[rc_id]['componenti'].append({
            'componentId': componente,
            'componentType': cod_tipo,
            'quantity': qty,
            'unitType': um if um else 'g'
        })

# Inserisci ricette nel database
for codice, data in ricette_dict.items():
    try:
        import json
        cursor.execute("""
            INSERT INTO final_recipes (id, code, name, category, components, yieldPercentage,
                                     productionOperations, conservationMethod, maxConservationTime,
                                     serviceWastePercentage, serviceWastePerIngredient, totalCost)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            generate(size=21),
            codice,
            codice.replace('_', ' ').title(),
            'Altro',
            json.dumps(data['componenti']),
            '95.0',
            json.dumps([]),
            'Refrigerato',
            '3 giorni',
            '5.0',
            json.dumps([]),
            '0.0'
        ))
        ricette_importate += 1
    except mysql.connector.IntegrityError:
        pass
    except Exception as e:
        print(f"Errore importando ricetta {codice}: {e}")

conn.commit()
print(f"Importate {ricette_importate} ricette finali")

# Chiudi connessione
cursor.close()
conn.close()

print("\n=== Importazione completata ===")
print(f"Totale ingredienti: {ingredienti_importati}")
print(f"Totale semilavorati: {semilavorati_importati}")
print(f"Totale ricette finali: {ricette_importate}")
