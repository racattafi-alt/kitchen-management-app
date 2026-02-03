import openpyxl
import os
import json
from collections import defaultdict
from urllib.parse import urlparse
import mysql.connector

# Carica il file Excel
wb = openpyxl.load_workbook('/home/ubuntu/upload/MENUUNION2026.xlsx', data_only=True)
ws_prezzo = wb['PREZZO_FINALE']
ws_rec = wb['REC_FINALE']

# Connessione database
db_url = os.environ.get('DATABASE_URL', '')
parsed = urlparse(db_url)

conn = mysql.connector.connect(
    host=parsed.hostname,
    port=parsed.port or 3306,
    user=parsed.username,
    password=parsed.password,
    database=parsed.path.lstrip('/'),
    ssl_ca='',
    ssl_verify_cert=False
)
cursor = conn.cursor(dictionary=True)

# Mappa ingredienti per nome (case-insensitive)
cursor.execute("SELECT id, name FROM ingredients")
ingredients_map = {}
for row in cursor.fetchall():
    ingredients_map[row['name'].lower()] = {'id': row['id'], 'name': row['name']}

# Mappa semilavorati per codice E per nome (case-insensitive)
cursor.execute("SELECT id, code, name FROM semi_finished_recipes")
semi_map_by_code = {}
semi_map_by_name = {}
for row in cursor.fetchall():
    semi_map_by_code[row['code']] = {'id': row['id'], 'name': row['name'], 'code': row['code']}
    semi_map_by_name[row['name'].lower()] = {'id': row['id'], 'name': row['name'], 'code': row['code']}

print(f"Ingredienti disponibili: {len(ingredients_map)}")
print(f"Semilavorati disponibili: {len(semi_map_by_code)}")

# Leggi quantità di output (batch size) dal foglio PREZZO_FINALE
batch_sizes = {}
headers_prezzo = [cell.value for cell in ws_prezzo[1]]
qty_pronto_idx = headers_prezzo.index('Quantità di prodotto pronto')
codice_idx = headers_prezzo.index('CODICE_RICETTA')

for row in ws_prezzo.iter_rows(min_row=2, values_only=True):
    if not row or not row[codice_idx]:
        continue
    codice = str(row[codice_idx]).strip()
    qty_pronto = row[qty_pronto_idx]
    if qty_pronto:
        # Converti in kg
        batch_sizes[codice] = float(qty_pronto) / 1000

print(f"\nBatch sizes trovati: {len(batch_sizes)}")

# Lista di operazioni da ignorare
OPERATIONS = ['cuoco', 'forno', 'mixer', 'piastra', 'abbattitore', 'tagliaverdure', 'friggitrice', 'arrotondatrice', 'impastatrice', 'lievitatore', 'macinacarne']

# Raggruppa componenti per RC_ID
recipes_data = defaultdict(list)
skipped_operations = set()
not_found = set()

for row_idx, row in enumerate(ws_rec.iter_rows(min_row=2, values_only=True), start=2):
    rc_id = row[0]  # Col A: RC_ID
    component_name = row[1]  # Col B: COMPONENTE
    cod = row[2]  # Col C: COD
    um = row[3]  # Col D: UM
    price_um = row[4]  # Col E: Prezzo_UM
    qty = row[5]  # Col F: Qty
    eur = row[6]  # Col G: EUR
    
    if not rc_id or not component_name:
        continue
    
    # Ignora operazioni
    if component_name.lower() in OPERATIONS or cod == 'ops':
        skipped_operations.add(component_name)
        continue
    
    # Determina se è ingrediente o semilavorato
    component_id = None
    component_type = None
    component_name_lower = component_name.lower()
    
    # 1. Cerca nei semilavorati per codice
    if cod and cod in semi_map_by_code:
        semi = semi_map_by_code[cod]
        component_id = semi['id']
        component_type = "SEMI_FINISHED"
    # 2. Cerca nei semilavorati per nome (case-insensitive)
    elif component_name_lower in semi_map_by_name:
        semi = semi_map_by_name[component_name_lower]
        component_id = semi['id']
        component_type = "SEMI_FINISHED"
    # 3. Cerca negli ingredienti per nome (case-insensitive)
    elif component_name_lower in ingredients_map:
        ing = ingredients_map[component_name_lower]
        component_id = ing['id']
        component_type = "INGREDIENT"
    else:
        not_found.add(f"{component_name} (cod: {cod})")
        continue
    
    # Converti quantità in grammi
    qty_g = float(qty) if qty else 0
    if um and um.lower() == 'kg':
        qty_g *= 1000
    
    component = {
        "type": component_type,
        "componentId": component_id,
        "quantity": qty_g,
        "unit": "k",
        "wastePercentage": 0,
    }
    
    recipes_data[rc_id].append(component)

print(f"\n=== STATISTICHE ===")
print(f"Ricette trovate: {len(recipes_data)}")
print(f"Operazioni ignorate: {len(skipped_operations)}")
print(f"Componenti non trovati: {len(not_found)}")

# Normalizza le quantità e aggiorna il database
print(f"\n=== NORMALIZZAZIONE E AGGIORNAMENTO DATABASE ===")
updated_count = 0
for rc_id, components in recipes_data.items():
    # Trova il batch size per questa ricetta
    batch_size_kg = batch_sizes.get(rc_id)
    
    if not batch_size_kg:
        print(f"⚠️  {rc_id}: Batch size non trovato, uso quantità originali")
        normalized_components = components
    else:
        # Normalizza le quantità dividendo per il batch size
        normalized_components = []
        for comp in components:
            normalized_comp = comp.copy()
            # Quantità originale in grammi / batch size in kg = grammi per kg
            normalized_comp['quantity'] = comp['quantity'] / batch_size_kg
            normalized_components.append(normalized_comp)
        
        print(f"✓ {rc_id}: Normalizzato per batch size {batch_size_kg:.2f} kg")
    
    components_json = json.dumps(normalized_components)
    
    # Trova la ricetta per codice
    cursor.execute("SELECT id, name FROM final_recipes WHERE code = %s", (rc_id,))
    recipe = cursor.fetchone()
    
    if recipe:
        cursor.execute(
            "UPDATE final_recipes SET components = %s WHERE id = %s",
            (components_json, recipe['id'])
        )
        updated_count += 1
    else:
        print(f"✗ Ricetta non trovata nel DB: {rc_id}")

conn.commit()
cursor.close()
conn.close()

print(f"\n=== COMPLETATO ===")
print(f"Ricette aggiornate: {updated_count}/{len(recipes_data)}")
